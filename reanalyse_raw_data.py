#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Raw-data audit for the BiB wine revision.

This script reads the original study workbooks used for revision verification:

* Results.xlsx: main-trial pressure/temperature workbook.
* Stack testing high temperature.xlsx: verification-trial workbook.

It writes derived CSV summaries and regenerated figure PNGs. It never modifies the
source workbooks or the CSV export.

Revision note:
Figures 3 and 4 now plot baseline-referred pressure change (ΔP) in the pressure
panels, rather than absolute pressure, to match the manuscript pressure endpoint
definition.
"""

from __future__ import annotations

import math
import shutil
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np
import openpyxl
import pandas as pd


MAIN_XLSX = Path("Results.xlsx")
VERIFICATION_XLSX = Path("Stack testing high temperature.xlsx")

MAIN_STORAGE_END_D = 20.0
HANDLING_END_D = 0.21
DAY20_START_D = 19.5

MAIN_SENSORS = [
    ("V50", "Top", "50", 2, 3),
    ("V53", "Top", "50", 4, 5),
    ("V71", "Top", "50", 6, 7),
    ("V51", "Bottom", "19", 8, 9),
    ("V59", "Bottom", "19", 10, 11),
    ("V52", "Bottom", "50", 12, 13),
    ("V56", "Bottom", "50", 14, 15),
    ("V70", "Bottom", "50", 16, 17),
    ("V55", "Top", "19", 18, 19),
    ("V57", "Top", "19", 20, 21),
    ("V58", "Top", "19", 22, 23),
]

VERIFICATION_SENSORS = [
    ("V26", "Top", "V26_TOP_T", "V26_TOP_P"),
    ("V62", "Top", "V62_TOP_T", "V62_TOP_P"),
    ("V27", "Bottom", "V27_BOT_T", "V27_BOT_P"),
    ("V64", "Bottom", "V64_BOT_T", "V64_BOT_P"),
]


def f_survival(f_value: float, df_num: int, df_den: int) -> float:
    if not np.isfinite(f_value) or f_value <= 0:
        return 1.0
    a = df_num / 2.0
    b = df_den / 2.0
    log_beta = math.lgamma(a) + math.lgamma(b) - math.lgamma(a + b)
    scale = df_num / df_den

    def pdf(x: float) -> float:
        if x <= 0:
            return 0.0
        return math.exp(
            a * math.log(scale)
            + (a - 1.0) * math.log(x)
            - (a + b) * math.log1p(scale * x)
            - log_beta
        )

    def simpson(left: float, right: float) -> float:
        mid = (left + right) / 2.0
        return (right - left) * (pdf(left) + 4.0 * pdf(mid) + pdf(right)) / 6.0

    def adaptive(left: float, right: float, eps: float, whole: float, depth: int) -> float:
        mid = (left + right) / 2.0
        left_area = simpson(left, mid)
        right_area = simpson(mid, right)
        if depth <= 0 or abs(left_area + right_area - whole) <= 15.0 * eps:
            return left_area + right_area + (left_area + right_area - whole) / 15.0
        return adaptive(left, mid, eps / 2.0, left_area, depth - 1) + adaptive(
            mid, right, eps / 2.0, right_area, depth - 1
        )

    cdf = adaptive(0.0, float(f_value), 1e-10, simpson(0.0, float(f_value)), 30)
    return max(0.0, min(1.0, 1.0 - cdf))


def t975(df: int) -> float:
    values = {1: 12.7062, 2: 4.3027, 3: 3.1824, 4: 2.7764, 5: 2.5706,
              6: 2.4469, 7: 2.3646, 8: 2.3060, 9: 2.2622, 10: 2.2281,
              11: 2.2010, 12: 2.1788}
    return values.get(int(df), 1.96)


def fit_lm(data: pd.DataFrame, response: str, terms: list[str]):
    cols = [np.ones(len(data))]
    if "position" in terms:
        cols.append((data["position"].values == "Top").astype(float))
    if "chamber" in terms:
        cols.append((data["chamber"].astype(str).values == "50").astype(float))
    if "interaction" in terms:
        cols.append(((data["position"].values == "Top") &
                     (data["chamber"].astype(str).values == "50")).astype(float))
    x = np.column_stack(cols)
    y = data[response].to_numpy(float)
    beta = np.linalg.lstsq(x, y, rcond=None)[0]
    residuals = y - x @ beta
    rank = np.linalg.matrix_rank(x)
    sse = float(residuals @ residuals)
    return beta, x, residuals, sse, len(y) - rank


def type2_anova(data: pd.DataFrame, response: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    beta, x, _resid, sse, df = fit_lm(data, response, ["position", "chamber"])
    _, _, _, sse_no_position, _ = fit_lm(data, response, ["chamber"])
    _, _, _, sse_no_chamber, _ = fit_lm(data, response, ["position"])
    _, _, _, sse_full, df_full = fit_lm(data, response, ["position", "chamber", "interaction"])
    mse = sse / df
    rows = []
    for factor, reduced_sse in [("position", sse_no_position), ("chamber", sse_no_chamber)]:
        ss = reduced_sse - sse
        f_value = ss / mse
        rows.append({"response": response, "factor": factor, "sum_sq": ss,
                     "df_num": 1, "df_den": df, "F": f_value,
                     "p": f_survival(f_value, 1, df)})
    f_interaction = ((sse - sse_full) / 1.0) / (sse_full / df_full)
    rows.append({"response": response, "factor": "position:chamber",
                 "sum_sq": sse - sse_full, "df_num": 1, "df_den": df_full,
                 "F": f_interaction, "p": f_survival(f_interaction, 1, df_full)})
    xtx_inv = np.linalg.inv(x.T @ x)
    tcrit = t975(df)
    contrasts = []
    for label, c in [("bottom_minus_top", np.array([0.0, -1.0, 0.0])),
                     ("nominal_50_minus_19", np.array([0.0, 0.0, 1.0]))]:
        estimate = float(c @ beta)
        se = math.sqrt(float((sse / df) * c @ xtx_inv @ c))
        contrasts.append({"response": response, "contrast": label,
                          "estimate_mbar": estimate,
                          "ci95_low_mbar": estimate - tcrit * se,
                          "ci95_high_mbar": estimate + tcrit * se,
                          "df": df})
    return pd.DataFrame(rows), pd.DataFrame(contrasts)


def load_main_trial() -> pd.DataFrame:
    wb = openpyxl.load_workbook(MAIN_XLSX, data_only=True, read_only=False)
    ws = wb["Compare"]
    cols = ["time_days", "time_h"] + [f"c{i}" for i in range(3, 25)]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 25)]
        if row[0] is None and row[1] is None:
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=cols)
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def main_sensor_map() -> pd.DataFrame:
    return pd.DataFrame([{"sensor": sid, "position": position,
                          "nominal_chamber_degC": chamber,
                          "temperature_column_zero_based": t_col,
                          "pressure_column_zero_based": p_col,
                          "stack_id": "not available",
                          "stack_height_cm": "not available",
                          "source_file": MAIN_XLSX.name,
                          "source_sheet": "Compare"}
                         for sid, position, chamber, t_col, p_col in MAIN_SENSORS])


def summarize_main_pressure(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    t_all = df["time_days"].to_numpy(float)
    valid = np.isfinite(t_all)
    t = t_all[valid]
    storage = t <= MAIN_STORAGE_END_D
    t = t[storage]
    rows = []
    early_rows = []
    for sid, position, chamber, t_col, p_col in MAIN_SENSORS:
        temp = df.iloc[:, t_col].to_numpy(float)[valid][storage]
        pressure = df.iloc[:, p_col].to_numpy(float)[valid][storage]
        p0 = float(pressure[np.isfinite(pressure)][0])
        delta_p = pressure - p0
        post_mask = t >= HANDLING_END_D
        full_mask = t >= 0.0
        early_mask = t <= (5.0 / 24.0)
        day20_mask = t >= DAY20_START_D
        post_idx = int(np.nanargmax(np.where(post_mask, delta_p, -np.inf)))
        full_idx = int(np.nanargmax(np.where(full_mask, delta_p, -np.inf)))
        early_idx = int(np.nanargmax(np.where(early_mask, delta_p, -np.inf)))
        rows.append({"sensor": sid, "position": position, "chamber": chamber,
                     "baseline_mbar": p0,
                     "baseline_rule": "initial valid pressure reading at time zero",
                     "peak_dP_mbar": float(delta_p[post_idx]),
                     "peak_day": float(t[post_idx]),
                     "peak_hour": float(t[post_idx] * 24.0),
                     "dP_day20_mbar": float(np.nanmean(delta_p[day20_mask])),
                     "day20_n_readings": int(np.sum(day20_mask))})
        early_rows.append({"sensor": sid, "position": position, "chamber": chamber,
                           "early_peak_dP_mbar": float(delta_p[early_idx]),
                           "early_peak_day": float(t[early_idx]),
                           "early_peak_hour": float(t[early_idx] * 24.0),
                           "full_record_peak_dP_mbar": float(delta_p[full_idx]),
                           "full_record_peak_day": float(t[full_idx]),
                           "full_record_peak_hour": float(t[full_idx] * 24.0),
                           "post_handling_peak_dP_mbar": float(delta_p[post_idx]),
                           "post_handling_peak_day": float(t[post_idx]),
                           "post_handling_peak_hour": float(t[post_idx] * 24.0),
                           "full_peak_equals_primary": bool(np.isclose(delta_p[full_idx], delta_p[post_idx])),
                           "early_min_temperature_degC": float(np.nanmin(temp[early_mask])),
                           "early_max_temperature_degC": float(np.nanmax(temp[early_mask]))})
    summary = pd.DataFrame(rows)
    early = pd.DataFrame(early_rows)
    anova_rows = []
    contrast_rows = []
    for response in ["peak_dP_mbar", "dP_day20_mbar"]:
        a, c = type2_anova(summary.rename(columns={response: "response_value"}), "response_value")
        a["original_response"] = response
        c["original_response"] = response
        anova_rows.append(a)
        contrast_rows.append(c)
    return summary, early, pd.concat(anova_rows, ignore_index=True), pd.concat(contrast_rows, ignore_index=True)


def summarize_main_temperature(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    t_all = df["time_days"].to_numpy(float)
    valid = np.isfinite(t_all)
    t = t_all[valid]
    storage = t <= MAIN_STORAGE_END_D
    t = t[storage]
    rows = []
    for sid, position, chamber, t_col, _p_col in MAIN_SENSORS:
        temp = df.iloc[:, t_col].to_numpy(float)[valid][storage]
        rows.append({"experiment": "main", "sensor": sid, "position": position,
                     "nominal_chamber_degC": chamber,
                     "initial_temperature_degC": float(temp[0]),
                     "temperature_5h_degC": float(temp[np.argmin(np.abs(t - 5.0 / 24.0))]),
                     "temperature_16h_degC": float(temp[np.argmin(np.abs(t - 16.0 / 24.0))]),
                     "post_handling_max_temperature_degC": float(np.nanmax(temp[t >= HANDLING_END_D])),
                     "post_handling_max_day": float(t[t >= HANDLING_END_D][np.nanargmax(temp[t >= HANDLING_END_D])]),
                     "late_mean_temperature_degC": float(np.nanmean(temp[t >= DAY20_START_D]))})
    per_sensor = pd.DataFrame(rows)
    group = (per_sensor.groupby(["experiment", "nominal_chamber_degC", "position"])
             .agg(n=("sensor", "count"),
                  initial_temperature_mean_degC=("initial_temperature_degC", "mean"),
                  initial_temperature_sd_degC=("initial_temperature_degC", "std"),
                  temperature_5h_mean_degC=("temperature_5h_degC", "mean"),
                  temperature_5h_sd_degC=("temperature_5h_degC", "std"),
                  temperature_16h_mean_degC=("temperature_16h_degC", "mean"),
                  temperature_16h_sd_degC=("temperature_16h_degC", "std"),
                  post_handling_max_temperature_mean_degC=("post_handling_max_temperature_degC", "mean"),
                  post_handling_max_temperature_sd_degC=("post_handling_max_temperature_degC", "std"),
                  late_mean_temperature_mean_degC=("late_mean_temperature_degC", "mean"),
                  late_mean_temperature_sd_degC=("late_mean_temperature_degC", "std"))
             .reset_index())
    return per_sensor, group


def load_verification_compare() -> pd.DataFrame:
    wb = openpyxl.load_workbook(VERIFICATION_XLSX, data_only=True, read_only=False)
    ws = wb["Compare"]
    headers = [ws.cell(1, col_idx).value if ws.cell(1, col_idx).value is not None else f"blank{col_idx}"
               for col_idx in range(1, 15)]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 15)]
        if all(value is None for value in row):
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=headers)
    for col in headers[1:]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def verification_sensor_map() -> pd.DataFrame:
    return pd.DataFrame([
        {"sensor": "V26", "position": "Top", "source_temperature_column": "V26_TOP_T",
         "source_pressure_column": "V26_TOP_P", "power_on": "2024-12-16 16:10",
         "stack": "single high-temperature three-box stack"},
        {"sensor": "V27", "position": "Bottom", "source_temperature_column": "V27_BOT_T",
         "source_pressure_column": "V27_BOT_P", "power_on": "2024-12-16 16:10",
         "stack": "single high-temperature three-box stack"},
        {"sensor": "V62", "position": "Top", "source_temperature_column": "V62_TOP_T",
         "source_pressure_column": "V62_TOP_P", "power_on": "2024-12-16 16:20",
         "stack": "single high-temperature three-box stack"},
        {"sensor": "V64", "position": "Bottom", "source_temperature_column": "V64_BOT_T",
         "source_pressure_column": "V64_BOT_P", "power_on": "2024-12-16 16:20",
         "stack": "single high-temperature three-box stack"},
    ])


def summarize_verification(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for sid, position, t_col, p_col in VERIFICATION_SENSORS:
        t = df["Time(days)"]
        temp = df[t_col]
        pressure = df[p_col]
        mask = t.notna() & temp.notna() & pressure.notna() & (t <= 15.0)
        first_idx = np.where(mask)[0][0]
        p0 = float(pressure.iloc[first_idx])
        delta_p = pressure - p0
        peak_idx = delta_p[mask].idxmax()
        min_idx = delta_p[mask].idxmin()
        temp_max_idx = temp[mask].idxmax()
        late = mask & (t >= 14.0)
        rows.append({"sensor": sid, "position": position,
                     "baseline_rule": "first valid pressure in verification Compare sheet",
                     "n_readings_to_day15": int(mask.sum()),
                     "first_day": float(t.iloc[first_idx]),
                     "last_day_to_day15": float(t[mask].max()),
                     "initial_temperature_degC": float(temp.iloc[first_idx]),
                     "maximum_temperature_degC": float(temp[mask].max()),
                     "maximum_temperature_day": float(t.loc[temp_max_idx]),
                     "temperature_day4_degC": float(temp[mask & (np.abs(t - 4.0) < (1.0 / 1440.0))].mean()),
                     "late_mean_temperature_day14_15_degC": float(temp[late].mean()) if late.any() else np.nan,
                     "baseline_pressure_mbar": p0,
                     "peak_dP_mbar": float(delta_p.loc[peak_idx]),
                     "peak_dP_day": float(t.loc[peak_idx]),
                     "peak_dP_hour": float(t.loc[peak_idx] * 24.0),
                     "minimum_dP_mbar": float(delta_p.loc[min_idx]),
                     "minimum_dP_day": float(t.loc[min_idx]),
                     "late_mean_dP_day14_15_mbar": float(delta_p[late].mean()) if late.any() else np.nan})
    return pd.DataFrame(rows)


def main_series(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    t = df["time_days"]
    for sid, position, chamber, t_col, p_col in MAIN_SENSORS:
        pressure = pd.to_numeric(df.iloc[:, p_col], errors="coerce")
        valid_pressure = pressure[np.isfinite(pressure)]
        if valid_pressure.empty:
            raise ValueError(f"No valid pressure readings found for sensor {sid}")
        p0 = float(valid_pressure.iloc[0])
        rows.append(pd.DataFrame({"time_days": t, "sensor": sid, "position": position,
                                  "chamber": chamber, "temperature_degC": df.iloc[:, t_col],
                                  "pressure_mbar": pressure, "delta_p_mbar": pressure - p0}))
    out = pd.concat(rows, ignore_index=True)
    out = out[out["time_days"].notna() & (out["time_days"] <= MAIN_STORAGE_END_D)]
    return out


def style_axis(ax):
    ax.grid(True, color="#dddddd", linewidth=0.8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)


def set_day_ticks(ax, ticks):
    ax.set_xticks(ticks)
    ax.set_xticklabels([str(int(t)) for t in ticks])


def set_hour_ticks(ax, ticks):
    ax.set_xticks(ticks)
    ax.set_xticklabels([str(int(t)) for t in ticks])


def add_phase_arrows(ax):
    trans = ax.get_xaxis_transform()
    arrow = dict(arrowstyle="<->", color="#777777", linewidth=1.0, shrinkA=0, shrinkB=0)
    ax.annotate("pressure rise", xy=(0.35, 0.92), xytext=(4.4, 0.92),
                xycoords=trans, textcoords=trans, ha="center", va="bottom",
                color="#777777", fontsize=9, arrowprops=arrow)
    ax.annotate("later decline", xy=(5.0, 0.92), xytext=(19.6, 0.92),
                xycoords=trans, textcoords=trans, ha="center", va="bottom",
                color="#777777", fontsize=9, arrowprops=arrow)


def plot_mean_sd(ax, series: pd.DataFrame, value: str, chamber: str, position: str,
                 color: str, label: str, x_col: str = "time_days",
                 marker: str | None = None, markersize: float = 0.0):
    subset = series[(series["chamber"] == chamber) & (series["position"] == position)]
    grouped = subset.groupby(x_col)[value].agg(["mean", "std"]).reset_index()
    x = grouped[x_col].to_numpy(float)
    mean = grouped["mean"].to_numpy(float)
    sd = grouped["std"].fillna(0).to_numpy(float)
    ax.plot(x, mean, color=color, linewidth=2.2, label=label, marker=marker, markersize=markersize)
    ax.fill_between(x, mean - sd, mean + sd, color=color, alpha=0.2)


def copy_final_figure(filename: str) -> None:
    package_dir = Path("SUBMISSION_PACKAGE") / "Figures"
    if package_dir.exists():
        shutil.copy2(filename, package_dir / filename)


def regenerate_figures(main_df: pd.DataFrame, pressure_summary: pd.DataFrame, verification_df: pd.DataFrame):
    series = main_series(main_df)
    amber = "#d28a00"
    blue = "#2f7ed8"
    green = "#21aa78"
    purple = "#5142ad"
    red = "#ef4b4b"

    fig, axes = plt.subplots(2, 2, figsize=(15, 9.8), sharex=True)
    for col, chamber, title in [(0, "50", "(A) C50 chamber set-point group"),
                                (1, "19", "(B) C19 chamber set-point group")]:
        axp = axes[0, col]
        axt = axes[1, col]
        plot_mean_sd(axp, series, "delta_p_mbar", chamber, "Top", amber, "Top BiB")
        plot_mean_sd(axp, series, "delta_p_mbar", chamber, "Bottom", blue, "Bottom BiB")
        plot_mean_sd(axt, series, "temperature_degC", chamber, "Top", amber, "Top BiB")
        plot_mean_sd(axt, series, "temperature_degC", chamber, "Bottom", blue, "Bottom BiB")
        axp.axhline(0, color="#777777", linestyle="--", linewidth=1.2)
        axt.axhline(float(chamber), color="#c7352d", linestyle="--", linewidth=1.5)
        axp.axvspan(0, 5.0 / 24.0, color="#eeeeee", alpha=0.9)
        axt.axvspan(0, 5.0 / 24.0, color="#eeeeee", alpha=0.9)
        add_phase_arrows(axp)
        axp.text(0.055, 0.18, "handling\n0-5 h", transform=axp.transAxes,
                 color="#555555", fontsize=9, va="bottom",
                 bbox=dict(facecolor="white", edgecolor="none", alpha=0.85, pad=1.5))
        axt.text(19.7, float(chamber) + 0.25, f"{chamber} °C set point",
                 color="#c7352d", fontsize=9, ha="right", va="bottom")
        axp.set_title(title, loc="left", fontsize=13, fontweight="bold")
        axp.set_ylabel("ΔP (mbar)")
        axt.set_ylabel("Temperature (°C)")
        axt.set_xlabel("Time after start (days)")
        set_day_ticks(axt, [0, 5, 10, 15, 20])
        style_axis(axp)
        style_axis(axt)
        axp.set_xlim(-0.4, 20.4)
    axes[0, 0].legend(loc="upper right", frameon=False)
    fig.tight_layout()
    fig.savefig("Figure3_pressure_temperature_FINAL.png", dpi=300)
    plt.close(fig)
    copy_final_figure("Figure3_pressure_temperature_FINAL.png")

    fig, axes = plt.subplots(2, 2, figsize=(15, 9.8), sharex=True)
    early = series[series["time_days"] <= (16.0 / 24.0)].copy()
    early["time_h"] = early["time_days"] * 24.0
    for col, chamber, title in [(0, "50", "(A) C50 chamber set-point group"),
                                (1, "19", "(B) C19 chamber set-point group")]:
        for row, value, ylabel in [(0, "delta_p_mbar", "ΔP (mbar)"),
                                   (1, "temperature_degC", "Temperature (°C)")]:
            ax = axes[row, col]
            for pos, color, label in [("Top", amber, "Top BiB"), ("Bottom", blue, "Bottom BiB")]:
                plot_mean_sd(ax, early, value, chamber, pos, color, label,
                             x_col="time_h", marker="o", markersize=3.5)
            ax.axvspan(0, 5, color="#eeeeee", alpha=0.9)
            if row == 0:
                ax.axhline(0, color="#777777", linestyle="--", linewidth=1.2)
            ax.set_ylabel(ylabel)
            style_axis(ax)
            set_hour_ticks(ax, [0, 4, 8, 12, 16])
        axes[0, col].set_title(title, loc="left", fontsize=13, fontweight="bold")
        axes[1, col].set_xlabel("Time from start of monitoring (h)")
        axes[0, col].set_xlim(-0.5, 16.5)
    axes[0, 0].annotate("transport and stacking\n(first ~5 h)",
                        xy=(5.0, 0.67), xytext=(8.0, 0.78),
                        xycoords=axes[0, 0].get_xaxis_transform(),
                        textcoords=axes[0, 0].get_xaxis_transform(),
                        arrowprops=dict(arrowstyle="->", color="#777777", linewidth=1.0),
                        color="#777777", fontsize=9, ha="left", va="center")
    axes[0, 0].legend(loc="upper right", frameon=False)
    fig.tight_layout()
    fig.savefig("Figure4_transient_16h_FINAL.png", dpi=300)
    plt.close(fig)
    copy_final_figure("Figure4_transient_16h_FINAL.png")

    fig, axes = plt.subplots(1, 2, figsize=(14, 6.4), sharey=False)
    responses = [("peak_dP_mbar", "(A) Post-handling peak ΔP"), ("dP_day20_mbar", "(B) Day-20 ΔP")]
    x_positions = {"19": 0, "50": 1}
    offsets = {"Top": -0.12, "Bottom": 0.12}
    for ax, (response, title) in zip(axes, responses):
        for chamber in ["19", "50"]:
            for position, color in [("Top", amber), ("Bottom", blue)]:
                vals = pressure_summary[(pressure_summary["chamber"] == chamber) &
                                        (pressure_summary["position"] == position)][response].to_numpy(float)
                x = x_positions[chamber] + offsets[position]
                ax.scatter(np.full(len(vals), x), vals, color=color, s=45, zorder=3)
                ax.errorbar(x, np.mean(vals), yerr=np.std(vals, ddof=1) if len(vals) > 1 else 0.0,
                            color=color, capsize=7, linewidth=2.2)
        ax.set_xticks([0, 1])
        ax.set_xticklabels(["C19", "C50"])
        ax.set_title(title, loc="left", fontsize=13, fontweight="bold")
        ax.set_ylabel("ΔP from own baseline (mbar)")
        style_axis(ax)
    axes[0].text(0.50, 0.96, "bottom > top at peak\nposition effect p=0.017",
                 transform=axes[0].transAxes, fontsize=10, ha="center", va="top",
                 bbox=dict(facecolor="white", edgecolor="#bbbbbb", boxstyle="round,pad=0.25", alpha=0.9))
    stat_box = dict(facecolor="white", edgecolor="none", alpha=0.86, pad=1.5)
    axes[0].text(0.22, 0.05, "position: F(1,8)=9.14, p=0.017\nchamber: F(1,8)=0.21, p=0.66",
                 transform=axes[0].transAxes, fontsize=10, bbox=stat_box)
    axes[1].text(0.04, 0.08, "position: F(1,8)=0.35, p=0.57\nchamber: F(1,8)=0.81, p=0.39",
                 transform=axes[1].transAxes, fontsize=10)
    axes[0].legend(handles=[Line2D([0], [0], marker="o", color=amber, label="Top BiB", linestyle="", markersize=7),
                            Line2D([0], [0], marker="o", color=blue, label="Bottom BiB", linestyle="", markersize=7)],
                   frameon=False, loc="upper right")
    fig.tight_layout()
    fig.savefig("Figure5_deltaP_FINAL.png", dpi=300)
    plt.close(fig)
    copy_final_figure("Figure5_deltaP_FINAL.png")

    fig, axes = plt.subplots(1, 2, figsize=(16.4, 7.0), sharey=True)
    label_box = dict(facecolor="white", edgecolor="none", alpha=0.88, pad=1.8)
    leader = dict(arrowstyle="-", color="#666666", linewidth=1.0)
    ax = axes[0]
    for pos, color, label in [("Top", amber, "C50 top"), ("Bottom", blue, "C50 bottom")]:
        plot_mean_sd(ax, series, "temperature_degC", "50", pos, color, label)
    ref = series[series["chamber"] == "19"].groupby("time_days")["temperature_degC"].agg(["mean", "std"]).reset_index()
    x = ref["time_days"].to_numpy(float)
    mean = ref["mean"].to_numpy(float)
    sd = ref["std"].fillna(0).to_numpy(float)
    ax.plot(x, mean, color="#777777", linewidth=2.0, label="C19 reference")
    ax.fill_between(x, mean - sd, mean + sd, color="#999999", alpha=0.2)
    ax.axhline(50, color="#c7352d", linestyle="--", linewidth=1.5)
    ax.set_title("(A) Main trial: C50 set point", loc="left", fontsize=13, fontweight="bold")
    ax.set_ylabel("Temperature inside the bag (°C)")
    ax.set_xlabel("Time after start (days)")
    ax.set_xlim(-0.4, 22.2)
    ax.set_ylim(18, 52.5)
    set_day_ticks(ax, [0, 5, 10, 15, 20])
    ax.text(21.7, 50.6, "50 °C set point", color="#c7352d", fontsize=9, ha="right", va="bottom", bbox=label_box)
    ax.annotate("top mean\n~34 °C", xy=(18.3, 34.3), xytext=(20.6, 38.0), color=amber,
                fontsize=10, fontweight="bold", ha="left", va="center", arrowprops=leader,
                bbox=label_box, annotation_clip=False)
    ax.annotate("bottom mean\n~26 °C", xy=(18.3, 26.2), xytext=(20.6, 29.0), color=blue,
                fontsize=10, fontweight="bold", ha="left", va="center", arrowprops=leader,
                bbox=label_box, annotation_clip=False)
    ax.annotate("", xy=(9.1, 50.0), xytext=(9.1, 34.8),
                arrowprops=dict(arrowstyle="<->", color="#c7352d", linewidth=1.2))
    ax.text(9.45, 42.4, "set-point shortfall\n(main trial)", color="#c7352d",
            fontsize=9, ha="left", va="center", bbox=label_box)
    ax.legend(frameon=False, loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=2, fontsize=9)
    style_axis(ax)
    ax = axes[1]
    for sid, position, t_col, _p_col in VERIFICATION_SENSORS:
        subset = verification_df[["Time(days)", t_col]].dropna()
        subset = subset[subset["Time(days)"] <= 15.0]
        color = {"V26": green, "V62": purple, "V27": red, "V64": amber}[sid]
        ax.plot(subset["Time(days)"], subset[t_col], color=color, linewidth=2.0, label=f"{sid} {position.lower()}")
    ax.axhline(50, color="#c7352d", linestyle="--", linewidth=1.5)
    ax.set_title("(B) Verification trial: C50 set point", loc="left", fontsize=13, fontweight="bold")
    ax.set_xlabel("Time after start (days)")
    ax.set_xlim(-0.4, 15.6)
    ax.set_ylim(18, 52.5)
    set_day_ticks(ax, [0, 3, 6, 9, 12, 15])
    ax.text(15.0, 50.6, "50 °C set point", color="#c7352d", fontsize=9, ha="right", va="bottom", bbox=label_box)
    ax.annotate("V26 stops\nafter ~4 days 8 h", xy=(4.35, 48.7), xytext=(4.9, 51.3),
                color="#555555", fontsize=9, ha="left", va="center", arrowprops=leader, bbox=label_box)
    ax.text(0.54, 0.42, "No in-bag sensor\nreached 50 °C", transform=ax.transAxes,
            color="#555555", fontsize=9, ha="center", va="center", bbox=label_box)
    ax.legend(frameon=False, loc="upper center", bbox_to_anchor=(0.5, -0.16), ncol=2, fontsize=9)
    style_axis(ax)
    fig.tight_layout(rect=(0, 0.16, 1, 1), w_pad=2.3)
    fig.savefig("Figure6_temperature_attained_FINAL.png", dpi=300)
    plt.close(fig)
    copy_final_figure("Figure6_temperature_attained_FINAL.png")

    fig, axes = plt.subplots(2, 1, figsize=(13.2, 9.4), sharex=True)
    for sid, position, t_col, p_col in VERIFICATION_SENSORS:
        subset = verification_df[["Time(days)", t_col, p_col]].dropna()
        subset = subset[subset["Time(days)"] <= 15.0]
        p0 = subset[p_col].iloc[0]
        color = {"V26": green, "V62": purple, "V27": red, "V64": amber}[sid]
        axes[0].plot(subset["Time(days)"], subset[p_col] - p0, color=color, linewidth=2.0, label=f"{sid} {position.lower()}")
        axes[1].plot(subset["Time(days)"], subset[t_col], color=color, linewidth=2.0, label=f"{sid} {position.lower()}")
    axes[0].axhline(0, color="#888888", linewidth=1.0)
    axes[1].axhline(50, color="#c7352d", linestyle="--", linewidth=1.5)
    axes[0].set_ylabel("ΔP from own baseline (mbar)")
    axes[1].set_ylabel("Temperature (°C)")
    axes[1].set_xlabel("Time after start (days)")
    axes[0].set_title("(A) Verification trial pressure", loc="left", fontsize=13, fontweight="bold")
    axes[1].set_title("(B) Verification trial temperature", loc="left", fontsize=13, fontweight="bold")
    for ax in axes:
        style_axis(ax)
        ax.set_xlim(-0.3, 15.3)
        set_day_ticks(ax, [0, 3, 6, 9, 12, 15])
    axes[0].annotate("V26 stops\nafter ~4 days 8 h", xy=(4.35, 5.0), xytext=(5.2, 20.0),
                     arrowprops=dict(arrowstyle="->", color="#777777", linewidth=1.0),
                     color="#555555", fontsize=9, ha="left",
                     bbox=dict(facecolor="white", edgecolor="#bbbbbb", boxstyle="round,pad=0.25", alpha=0.88))
    axes[0].annotate("common dip\naround day 6", xy=(6.2, -25.0), xytext=(7.8, -11.0),
                     arrowprops=dict(arrowstyle="->", color="#777777", linewidth=1.0),
                     color="#555555", fontsize=9, ha="left",
                     bbox=dict(facecolor="white", edgecolor="#bbbbbb", boxstyle="round,pad=0.25", alpha=0.88))
    axes[1].text(15.0, 50.5, "50 °C set point", color="#c7352d", fontsize=9, ha="right", va="bottom")
    axes[1].text(0.56, 0.50, "No in-bag sensor\nreached 50 °C",
                 transform=axes[1].transAxes, color="#555555", fontsize=9, ha="center", va="center",
                 bbox=dict(facecolor="white", edgecolor="none", alpha=0.88, pad=1.8))
    axes[0].legend(frameon=False, loc="upper right", ncol=4)
    fig.tight_layout()
    fig.savefig("Figure7_verification_trial_FINAL.png", dpi=300)
    plt.close(fig)
    copy_final_figure("Figure7_verification_trial_FINAL.png")


def main() -> None:
    main_df = load_main_trial()
    main_sensor_map().to_csv("main_trial_sensor_map.csv", index=False)
    pressure, early, anova, contrasts = summarize_main_pressure(main_df)
    pressure.to_csv("pressure_summary_raw.csv", index=False)
    early.to_csv("early_transient_sensitivity.csv", index=False)
    anova.to_csv("pressure_anova_raw.csv", index=False)
    contrasts.to_csv("pressure_contrasts_raw.csv", index=False)
    temp_sensor, temp_group = summarize_main_temperature(main_df)
    temp_sensor.to_csv("thermal_summary_main_sensors.csv", index=False)
    temp_group.to_csv("thermal_summary_main_groups.csv", index=False)
    verification_df = load_verification_compare()
    verification_sensor_map().to_csv("verification_trial_sensor_map.csv", index=False)
    verification = summarize_verification(verification_df)
    verification.to_csv("verification_summary_raw.csv", index=False)
    regenerate_figures(main_df, pressure, verification_df)
    print("Wrote derived CSV summaries and regenerated Figures 3-7 FINAL PNGs.")
    print("Figures 3 and 4 pressure panels now use baseline-referred ΔP.")
    print("Copied final figures into SUBMISSION_PACKAGE/Figures where that directory exists.")
    print("Main pressure rows:", len(pressure))
    print("Verification sensor rows:", len(verification))


if __name__ == "__main__":
    main()
